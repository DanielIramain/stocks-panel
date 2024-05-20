import requests
from tkinter.filedialog import asksaveasfile 

import pandas as pd
import openpyxl
import openpyxl.utils.dataframe

import view

def choose_service(service_choosen: str):
    URL = f'https://www.alphavantage.co/query?function={service_choosen}&symbol={view.simbol}&apikey={view.api_key}'

    return URL

def request_information():
    '''
    Extrae la información solicitada a través de la API
    Convertimos de JSON a dataframe y transponemos los datos para presentarlos
    '''
    from view import service

    global data
    global df
    
    URL = choose_service(service)

    request = requests.get(URL)
    data = request.json()
    df = pd.DataFrame.from_dict([data])
    df = pd.DataFrame.transpose(df)

    obtain_fundamentals()

def obtain_fundamentals():
    '''
    Pregunta el tipo de servicio y en funcion a ello presenta la informacion
    La estructura de control esta definida segun el servicio solicitado
    '''
    
    global quarterly_report
    
    if view.service == 'overview':
        try:
            with asksaveasfile(mode='w', defaultextension='.xlsx') as file:
                df.to_excel(file.name)
        except AttributeError:
            print("El usuario ha cancelado")
    else:
        '''
        Los datos en la variable data deben ser normalizados en un dataframe
        Normalizamos en un dataframe los datos en formato JSON y luego transponemos
        '''
        df_normalized = pd.json_normalize(data)
        df_normalized_transposed = pd.DataFrame.transpose(df_normalized)
        
        '''
        Transponemos el dataframe de los datos normalizados
        Accedemos a las ganancias o reportes cuatrimestrales
        '''
        if view.service == 'earnings':
            df_quarterly_report = df_normalized_transposed.loc['quarterlyEarnings']
            quarterly_report = df_quarterly_report.iloc[0]
        else:
            df_quarterly_report = df_normalized_transposed.loc['quarterlyReports']
            quarterly_report = df_quarterly_report.iloc[0]

        save_information()

def save_information() -> None:
        '''
        Creamos un workbook
        Seleccionamos la worksheet
        '''
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        '''
        Accedemos a cada reporte cuatrimestral
        Escribimos cada reporte (dataframe) en el worksheet
        ''' 
        for report in range(0, len(quarterly_report), 1):
            dic_report = quarterly_report[report]
            df_dic_report = pd.DataFrame([dic_report])
            
            '''
            Esta condicion evalua si la hoja de trabajo esta vacia para asignar el header
            de lo contrario inserta los datos
            '''
            if worksheet.max_row == 1:
                rows = openpyxl.utils.dataframe.dataframe_to_rows(df_dic_report, index=False, header=True)
            else:
                rows = openpyxl.utils.dataframe.dataframe_to_rows(df_dic_report, index=False, header=False)

            #Finalmente agrega cada fila a la worksheet
            for r in rows:
                worksheet.append(r)
        
        with asksaveasfile(mode='w', defaultextension='.xlsx') as file:
            workbook.save(file.name)

        print(df_dic_report)
        print(worksheet.max_row)