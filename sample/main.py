import requests
import pandas as pd
from tkinter.filedialog import asksaveasfile 

import openpyxl
import openpyxl.utils.dataframe

import view

def elegir_funcion(funcion: str):
    global URL
    
    URL = f'https://www.alphavantage.co/query?function={funcion}&symbol={view.simbolo}&apikey={view.API_KEY}'    

    #return funcion, URL

def solicitar_informacion():
    ###Extrae la información solicitada a través de la API y la presenta al cliente
    r = requests.get(URL)
    data = r.json()
    ###Convertimos a df y transponemos los datos del df para presentarlos
    df = pd.DataFrame.from_dict([data])
    df = pd.DataFrame.transpose(df)
    
    ###Pregunta el tipo de servicio y en funcion a eso presenta la informacion
    ###Para el caso de overview
    if view.servicio == 'overview':
        try:
            with asksaveasfile(mode='w', defaultextension='.xlsx') as file:
                df.to_excel(file.name)
        except AttributeError:
            print("The user canceled")
    else:
        df_normalizado = pd.json_normalize(data)
        df_normalizado_transpuesto = pd.DataFrame.transpose(df_normalizado)
        ###Para los demas casos:
        if view.servicio == 'earnings':
            df_reportes_cuatrimestrales = df_normalizado_transpuesto.loc['quarterlyEarnings']
            reportes_cuatrimestrales = df_reportes_cuatrimestrales.iloc[0]
        else:
            df_reportes_cuatrimestrales = df_normalizado_transpuesto.loc['quarterlyReports']
            reportes_cuatrimestrales = df_reportes_cuatrimestrales.iloc[0]

        ###Creamos un workbook
        workbook = openpyxl.Workbook()
        ###Seleccionamos la worksheet
        worksheet = workbook.active

        ###Escribimos cada df (primero lo transformamos) en worksheet 
        for reporte in range(0, len(reportes_cuatrimestrales), 1):
            dic_reporte = reportes_cuatrimestrales[reporte]
            
            df_dic_reporte = pd.DataFrame([dic_reporte])
            
            #Esta condicion evalua si la hoja de trabajo esta vacia para asignar el header
            #de lo contrario inserta los datos
            if worksheet.max_row == 1:
                rows = openpyxl.utils.dataframe.dataframe_to_rows(df_dic_reporte, index=False, header=True)
            else:
                rows = openpyxl.utils.dataframe.dataframe_to_rows(df_dic_reporte, index=False, header=False)

            for r in rows:
                worksheet.append(r)
        
        with asksaveasfile(mode='w', defaultextension='.xlsx') as file:
            workbook.save(file.name)

        print(df_dic_reporte)
        print(worksheet.max_row)