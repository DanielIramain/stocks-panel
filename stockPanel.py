#Imports
##Librerias basicas
import matplotlib.pyplot as plt
import pandas as pd
import csv
import openpyxl
import openpyxl.utils.dataframe

##Tkinter
from tkinter import *
from tkinter import ttk, messagebox

##Alpha Vantage
import alphaVintageConfig as av
from alpha_vantage.timeseries import TimeSeries

##Otros imports
import config

#Funciones globales
def capturar_datos():
    ###Se encarga de capturar los datos mostrados a través de GUI para ser usados en los métodos
    global simbolo
    global servicio
    
    simbolo = entrada_ticker.get()
    servicio = combo.get()
    
    print('simbolo: ', simbolo)
    print('servicio: ', servicio)
    
    return simbolo

def mostrar_tabla(ts, data):
    print(ts)
    print(type(ts))
    print(data)
    print(type(data))

def elegir_funcion(funcion: str):
    global URL
    
    URL = f'https://www.alphavantage.co/query?function={funcion}&symbol={simbolo}&apikey={av.API_KEY}'     

    return funcion, URL

def solicitar_informacion():
    ###Extrae la información solicitada a través de la API y la presenta al cliente
    r = av.requests.get(URL)
    
    data = r.json()
    
    df = pd.DataFrame.from_dict([data])
    ###Transponemos los datos del df para presentarlos
    df = pd.DataFrame.transpose(df)
    ###Pregunta el tipo de servicio y en funcion a eso presenta la informacion
    if servicio == 'overview':
        df.to_excel(r'D:\Documents\03_Programación y desarrollo\stocks-panel\prueba.xlsx', index=True)
    else:
        df_normalizado = pd.json_normalize(data)
        df_normalizado_transpuesto = pd.DataFrame.transpose(df_normalizado)
        if servicio == 'earnings':
            df_reportes_cuatrimestrales = df_normalizado_transpuesto.loc['quarterlyEarnings']
            reportes_cuatrimestrales = df_reportes_cuatrimestrales.iloc[0]
        else:
            df_reportes_cuatrimestrales = df_normalizado_transpuesto.loc['quarterlyReports']
            reportes_cuatrimestrales = df_reportes_cuatrimestrales.iloc[0]

        ###Creamos un workbook
        workbook = openpyxl.Workbook()
        ###Seleccionamos la worksheet
        worksheet = workbook.active

        ###Escribimos cada df (primero lo transformamos) en worksheet 
        for reporte in range(0, len(reportes_cuatrimestrales), 1):
            dic_reporte = reportes_cuatrimestrales[reporte]
            
            df_dic_reporte = pd.DataFrame([dic_reporte])
            
            #Esta condicion evalua si la hoja de trabajo esta vacia para asignar el header
            #de lo contrario inserta los datos
            if worksheet.max_row == 1:
                rows = openpyxl.utils.dataframe.dataframe_to_rows(df_dic_reporte, index=False, header=True)
            else:
                rows = openpyxl.utils.dataframe.dataframe_to_rows(df_dic_reporte, index=False, header=False)

            for r in rows:
                worksheet.append(r)

        workbook.save('prueba.xlsx')

        print(df_dic_reporte)
        print(worksheet.max_row)
        
def obtener_listado(funcion:str):
    global funcion_elegida
    global CSV_URL
    
    funcion_elegida = funcion
    CSV_URL = f'https://www.alphavantage.co/query?function={funcion}&apikey={av.API_KEY}'

    with av.requests.Session() as s:
        descarga = s.get(CSV_URL)
        decodificacion = descarga.content.decode('utf-8')
        cr = csv.reader(decodificacion.splitlines(), delimiter=',')
        listado = list(cr)
        ###Si quiero recorrer cada fila e imprimir como una lista cada elemento
        #for fila in listado:
            #print(fila)
        data = pd.DataFrame(listado)
        print(data)

#Clases
class Fundamentos():
    def __init__(self, simbolo: str):
        self.simbolo = simbolo

        return self

    def obtener_fundamentos():
        elegir_funcion(servicio)
        solicitar_informacion()

class DatosMercado():
    def __init__(self) -> None:
        return self
    
    def obtener_activos_vigentes():
        ###Obtener listado de activos en vigencia (imprime por pantalla un dataframe)
        obtener_listado('LISTING_STATUS')
    
    def obtener_calendario_ganancias(self):
        ###Obtener listado de companias que presentan ganancias en los proximos meses (3, 6 u 12) 
        CSV_URL = f'https://www.alphavantage.co/query?function=EARNINGS_CALENDAR&horizon=3month&apikey={av.API_KEY}'

        obtener_listado('EARNINGS_CALENDAR')
    
    def obtener_calendario_ipo(self):
        ###Obtener calendario de primera oferta publica en la bolsa (de EEUU)
        obtener_listado('IPO_CALENDAR')

class SeriesDeTiempo():
    def Intradia(simbolo, intervalo):
        if config.mercado == 'EEUU' and config.categoria == 'Historicos':
            ts = TimeSeries(key=av.API_KEY, output_format='pandas')
            data, meta_data = ts.get_intraday(symbol=simbolo, interval=intervalo, outputsize='full')
            mostrar_tabla(ts, data)
    def IntradiaExtendido(simbolo, intervalo):
        if config.mercado == 'EEUU' and config.categoria == 'Historicos':
            ts = TimeSeries(key=av.API_KEY, output_format='csv')
            data, meta_data = ts.get_intraday_extended(symbol=simbolo, interval=intervalo)
            df = pd.DataFrame(data)
            mostrar_tabla(ts, data)
            print(df)
            print(ts)
    def DiarioAjustado(simbolo):
        if config.mercado == 'EEUU' and config.categoria == 'Historicos':
            ts = TimeSeries(key=av.API_KEY, output_format='pandas')
            data, meta_data = ts.get_daily_adjusted(simbolo)
            mostrar_tabla(ts, data)

    #DiarioAjustado(simbolo)

class NoticiasAlpha():
    ###Buscar la forma de acceder a la información del JSON (diccionario) => documentación Alpha Vantage oficial
    def NoticiasMercado(simbolo):
        FUNCION = 'NEWS_SENTIMENT'
        url = av.URL = f'https://www.alphavantage.co/query?function={FUNCION}&tickers={simbolo}&apikey={av.API_KEY}'
        url_request = av.requests.get(url)
        data = url_request.json()

        print(data)
        print(type(data))

#GUI
root = Tk()
ticker = StringVar()
frame = ttk.Frame(root, padding=100)
combo = ttk.Combobox(state='readonly', 
                     values=['overview', 
                             'income_statement', 
                             'balance_sheet', 
                             'cash_flow',
                             'earnings'])


frame.grid()
ttk.Label(frame, text='Escriba el ticker').grid(column=0, row=0)

entrada_ticker = ttk.Entry(frame)
entrada_ticker.grid(column=1, row=0)

ttk.Button(frame, text='Mostrar datos', command=Fundamentos.obtener_fundamentos).grid(column=2, row=1)
ttk.Button(frame, text='Guardar datos', command=capturar_datos).grid(column=1, row=1)
combo.place(x=50, y=50)

root.mainloop()